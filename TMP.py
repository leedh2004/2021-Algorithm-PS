#-*- coding:utf-8 -*-
import openpyxl
import os
import pandas
import psycopg2
import time
import datetime
from datetime import timedelta
from tqdm import trange
import openpyxl

def url_parsing(url_rear_part,target_url):
    starting_point = target_url.find("http")
    if target_url.count("http") > 1:
        target_url = target_url[starting_point:target_url.find("http",starting_point+len("http"))]
        starting_point = target_url.find("http")
    elif target_url.find("http") < 0:
        starting_point = target_url.find("www")   
    if "storefarm.naver.com" in target_url:
        tmp = target_url.find("storefarm.naver.com")
        tmp2 = target_url.find('/',tmp+len("storefarm.naver.com")+1)
        tmp3 = target_url.find('?',tmp+len("storefarm.naver.com")+1)
        if tmp2 == -1 and tmp3==-1:
            return target_url[starting_point:]+"/"
        elif tmp2== -1 and tmp3!=-1:
            return target_url[starting_point:tmp3]+"/"
        else :
            return target_url[starting_point:tmp2+1]
    if "smartstore.naver.com" in target_url:
        tmp = target_url.find("smartstore.naver.com")
        tmp2 = target_url.find('/',tmp+len("smartstore.naver.com")+1)
        tmp3 = target_url.find('?',tmp+len("smartstore.naver.com")+1)
        if tmp2 == -1 and tmp3==-1:
            return target_url[starting_point:]+"/"
        elif tmp2== -1 and tmp3!=-1:
            return target_url[starting_point:tmp3]+"/"
        else :
            return target_url[starting_point:tmp2+1]
    else :
        for each_rear_part in url_rear_part:
            if each_rear_part in target_url:
                return target_url[starting_point:target_url.find(each_rear_part)+len(each_rear_part)]
        for each_rear_part in url_rear_part:
            form1 = each_rear_part[0:len(each_rear_part)-1] + ":"
            if form1 in target_url:
                return target_url[starting_point:target_url.find(form1)+len(form1)-1] + "/"
        for each_rear_part in url_rear_part:
            form2 = each_rear_part[0:len(each_rear_part)-1]
            if form2 in target_url:
                return target_url[starting_point:target_url.find(form2)+len(form2)] + "/"
        print ("안걸린애들"+target_url)
        if target_url[-1]=='/':
            if target_url[-2]=='/':
                 return target_url[starting_point:-1]   
            return target_url[starting_point:]
        else :
            return target_url[starting_point:]+"/"

def insert_sandbox(name,url,cate,score,now_time):
    try :
        #print ("INSERT INTO sb_pbs.tmp_plca_brand_manufacturer_url_score_test VALUES ('%s','%s','%s','%s','%s');"%(name,url,cate,score,now_time))
        cursor_sandbox.execute("INSERT INTO sb_pbs.tmp_plca_brand_manufacturer_url_score VALUES ('%s','%s','%s','%s','%s');"%(name,url,cate,score,now_time))
        conn_sandbox.commit()
    except Exception as e:
        print(e)
        #print ("INSERT INTO sb_pbs.tmp_plca_brand_manufacturer_url_score_test VALUES ('%s','%s','%s','%s','%s');"%(name,url,cate,score,now_time))
        
def update_sandbox(name,url,cate,now_time):
    try:
        #print("UPDATE sb_pbs.tmp_plca_brand_manufacturer_url_score_test SET score = score+1 WHERE name='%s' AND cate='%s' AND url='%s';"%(name,cate,url))
        cursor_sandbox.execute("UPDATE sb_pbs.tmp_plca_brand_manufacturer_url_score SET score = score+1 WHERE name='%s' AND cate='%s' AND url='%s';"%(name,cate,url))
        conn_sandbox.commit()
        #print ("UPDATE sb_pbs.tmp_plca_brand_manufacturer_url_score_test SET update_dy = '%s' WHERE name='%s' AND cate='%s' AND url='%s';"%(now_time,name,cate,url))
        cursor_sandbox.execute("UPDATE sb_pbs.tmp_plca_brand_manufacturer_url_score SET update_dy = '%s' WHERE name='%s' AND cate='%s' AND url='%s';"%(now_time,name,cate,url))
        conn_sandbox.commit()
    except Exception as e:
        print(e)

def update_date_sandbox(name,url,cate,now_time):
    try:
        #print ("UPDATE sb_pbs.tmp_plca_brand_manufacturer_url_score_test SET update_dy = '%s' WHERE name='%s' AND cate='%s' AND url='%s';"%(now_time,name,cate,url))
        cursor_sandbox.execute("UPDATE sb_pbs.tmp_plca_brand_manufacturer_url_score SET update_dy = '%s' WHERE name='%s' AND cate='%s' AND url='%s';"%(now_time,name,cate,url))
        conn_sandbox.commit()
    except Exception as e:
        print(e)   

def find_cate(skuid):
    try : 
        cursor_sandbox.execute("SELECT mngcateid FROM bimart.dwd_sku_x_pl_margin WHERE skuseq="+str(skuid)+";")
        output_sandbox = cursor_sandbox.fetchall()
        col_names_sandbox = [desc[0] for desc in cursor_sandbox.description]
        df_sandbox = pandas.DataFrame(output_sandbox, columns=col_names_sandbox)
        cursor_sandbox.execute("SELECT cate2 FROM bimart.management_category_hier_curr WHERE mngcateid="+str(df_sandbox.iloc[0,0])+";")
        output_sandbox = cursor_sandbox.fetchall()
        col_names_sandbox = [desc[0] for desc in cursor_sandbox.description]
        df_sandbox = pandas.DataFrame(output_sandbox, columns=col_names_sandbox)
        return str(df_sandbox.iloc[0,0])
    except Exception as e:
        print(e)
        
def check_is_there(name,url,cate):
    try:
        cursor_sandbox.execute("SELECT COUNT(*) FROM sb_pbs.tmp_plca_brand_manufacturer_url_score WHERE name='%s' AND cate='%s' AND url='%s';"%(name,cate,url))
        output_sandbox = cursor_sandbox.fetchall()
        col_names_sandbox = [desc[0] for desc in cursor_sandbox.description]
        df_sandbox = pandas.DataFrame(output_sandbox, columns=col_names_sandbox)
        if int(df_sandbox.iloc[0,0]) >0:
            return 1
        else :
            return 0
    except Exception as e:
        print(e)

def is_valid_url(word):
    if "http" in word :
        return True
    elif "www" in word:
        return True
    else:
        return False


#connect sandbox db
user_sandbox = "dmsghjang"
password_sandbox = "Dmsghjang1!"
dbname_sandbox = "sandbox"
host_sandbox = "dw-sandbox.coupang.net"
port_sandbox = 5439
conn_str_sandbox = "dbname='{}' user='{}' host='{}' password='{}' port={}".format(dbname_sandbox, user_sandbox, host_sandbox, password_sandbox, port_sandbox)
conn_sandbox = psycopg2.connect(conn_str_sandbox)
cursor_sandbox = conn_sandbox.cursor()

#now_time
now = time.localtime()
now_time = str("%04d%02d%02d" % (now.tm_year, now.tm_mon, now.tm_mday))
now_time_for_log="%04d%02d%02d%02d:%02d" % (now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour,now.tm_min)


#pandas_setup
pandas.options.display.max_columns = 9999999999999
pandas.options.display.max_rows = 99999999999999

#url_parsing_sop
url_rear_part = ['.co.kr/','.com/','.kr/' ,'.net/', '.jp/','.us/', '.au/', '.de/','.uk/','.fr/','.nl/','.re/','.cn/','.ee/','.life/','.pk/','.bio/','.do/','.fage/','.world/','.mu/','.as/','.gift/','.shop/','.se/','.co/','.cz/','.ro/','.mc/','.watch/','.sk/','.land/','.mx/','.ru/','.games/','.pe/','.pet/','.lt/','.asia/','.gov/','.fun/','.hr/','.pg/','.tv/','.at/','.it/','ie/','.hk/','.info/','.th/','.gr/','.eu/','.ge/','.si/','.ca/','.gr/','.biz/','.pt/','.si/','.es/','.ph/','.be/','.yt/','.fi/','.io/','.no/','.st/','.pro/','.vn/','.be/','.cc/','.ch/','.global/','.org/','.dk/','.so/','.fi/','.store/','online/','.direct/','.me/','.pl/','.tw/','.in/','.nz/','.ae/','bg/','edu/','.sg/','.lk/','.php/','.my/','.uk/','.top/','.cl/']
openmarket_url_list = ["gmarket.co.kr","11st.co.kr","walmart.com","auction.co.kr","interpark.com","aliexpress.com","g9.co.kr","kyobobook.co.kr","yes24.com","yes24.co.kr","lotte.com","ssg.com","hyundaihmall.com","shopping.naver.com","blog.naver.com","interpark.com","coupang.com","wemakeprice.com","tmon.co.kr","amazon.com","namu.wiki","map.naver","facebook.com","instagram.com","twitter.com","danawa.com"]
trash = ['','브랜드없음', '\x05',' ','-','중국','b', '0', '중국oem', '상세설명참조', 'made in china', '상세컨텐츠참조', '브랜드 없음','중국OEM', '중국 OEM', '제조사', '한국', '기타', '상세이미지참조', '상세페이지참조', '알수없음', '상품상세참조', '자체브랜드', '상세참조', '없음', '중국공장', '중국제조', 'OEM', '중국', '해외아시아중국', '한국제조', '업체미제공으로', '알수없음', '해당사항없음', '알수없음', '브래드없음', '상세컨테츠참조', '기타상세참조', '상세정보별도표기', '상세컨첸츠참조', '상품상세설명참조', '상품상세페이지참조', '상세페이지기재', '상세정보참조', '상세페이지참조', '상세페이지참고', '상세설명참조', '상세정보', '별도표기', '상세설명참고', '상세컨텐츠참고', '미국', '태국OEM']
status_list=['OBSOLETE_LOW_DEMAND', 'OBSOLETE_CODE_ERROR', '해당없음', 'OBSOLETE_HEAVY_BULKY', '']


#val
total_sku=0

sku_trash=1
sku_cover_mother=2
url_m_or_b=3
sku_cover_null_null=4
sku_precision_mother=5
sku_hit=6
sku_precision_null_null=7

m_trash=8
m_cover_mother=9
url_m=10
m_cover_null_null=11
m_precision_mother=12
m_hit=13
m_precision_null_null=14

b_trash=15
b_cover_mother=16
url_b=17
b_cover_null_null=18
b_precision_mother=19
b_hit=20
b_precision_null_null=21



cate_list = ['TOTAL','출산/유아', '식품', '가전디지털', '스포츠/레저용품', '패션잡화', '가구/홈인테리어', '생활용품', '뷰티', '문구/사무용품', '완구/취미', '자동차용품', '패션의류', '주방용품', '반려/애완용품', '도서/음반','로켓프레시']
list_of_var = ['total_sku', 'sku_trash', 'sku_cover_mother', 'url_m_or_b', 'sku_cover_null_null', 'sku_precision_mother', 'sku_hit', 'sku_precision_null_null', 'm_trash', 'm_cover_mother', 'url_m', 'm_cover_null_null', 'm_precision_mother', 'm_hit', 'm_precision_null_null', 'b_trash', 'b_cover_mother', 'url_b', 'b_cover_null_null', 'b_precision_mother', 'b_hit', 'b_precision_null_null']
list_of_columm=['end_date','skuid','skuname','manufacture','system_ulr_manu','brand_kr','system_ulr_brand','after_parsing_audit_url','audit_url','status','cate','Garbage','sys_url존재여부','op_url존재여부','일치여부']
data_list=[[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]]

start = time.time()
wb = openpyxl.Workbook()
ws = wb.active
ws.title='Data'
ws_list=[]

ws_sku=wb.create_sheet()
ws_sku.title="SKU"
sku_now_line=1


for t in range(len(list_of_columm)):
    ws_sku.cell(row = sku_now_line,column=t+1,value=str(list_of_columm[t]))
sku_now_line=sku_now_line+1
ws_sku_list=[]


#cursor_sandbox.execute("SELECT groupware_timestamp,skuid,retailcate_1,manufacture,system_ulr_manu,brand_kr,system_ulr_brand,url_manufacturer,detail_manufacturer,approval_po_status FROM sb_pbs.tmp_renault_oat_kanban_acc UNION SELECT groupware_timestamp,skuid,retailcate_1,manufacture,system_ulr_manu,brand_kr,system_ulr_brand,url_manufacturer,detail_manufacturer,approval_po_status FROM sb_pbs.tmp_renault_oat_kanban_now" )
print("커버율,정확도 계산에 필요한 데이터를 acc와 now 테이블에서 가져오는중입니다. 잠시만 기다려 주십시오...")
try:
    cursor_sandbox.execute('''
                            SELECT a.skuid,a.sku_nm,a.manufacture,a.system_ulr_manu,a.brand_kr,a.system_ulr_brand,a.url_manufacturer,a.audit_end_dt,a.approval_po_status,c.cate2
                            FROM
                            (SELECT skuid,sku_nm,manufacture,system_ulr_manu,brand_kr,system_ulr_brand,url_manufacturer,system_detail_manu,audit_end_dt,approval_po_status
                            FROM sb_pbs.tmp_renault_oat_kanban_acc
                            UNION
                            SELECT skuid,sku_nm,manufacture,system_ulr_manu,brand_kr,system_ulr_brand,url_manufacturer,system_detail_manu,audit_end_dt,approval_po_status
                            FROM sb_pbs.tmp_renault_oat_kanban_now
                            ORDER BY audit_end_dt
                            )a
                            JOIN  bimart.dwd_sku_x_pl_margin b on a.skuid=b.skuseq
                            JOIN bimart.management_category_hier_curr c on c.mngcateid=b.mngcateid
                            ''')
    output_sandbox = cursor_sandbox.fetchall()
    col_names_sandbox = [desc[0] for desc in cursor_sandbox.description]
    df_sandbox = pandas.DataFrame(output_sandbox, columns=col_names_sandbox)
    print("1.주차 데이터를 원하시면 년도와주를 입력해주십시오.ex)2019년도48주차-201948, 2019년도1주차-201901")
    print("2.시작날짜와 종료날짜로  시작일8자리와 ~ 종료일 8자리를 입력해주십시오.ex)20191201~20191211")
    answer = input("입력 : ")
    if  '~' in answer:
        start_date,end_date=answer.split('~')
    else :
        now_year = answer[0:4]
        now_week = answer[4:]
    for i in range(len(df_sandbox)):
        #쓰레기값인지 아닌지 파악해주는 변수
        check_valid_manufacture=1
        check_valid_brand_kr=1

        #엔드데이트가 안찍혀 있는애들 걸러줘
        try:
            if len(str(int(df_sandbox.iloc[i,7]))) !=8:
                continue
        except Exception as e:
            continue
        year=int(str(int(df_sandbox.iloc[i,7]))[0:4])
        month=int(str(int(df_sandbox.iloc[i,7]))[4:6])
        day=int(str(int(df_sandbox.iloc[i,7]))[6:8])
        #해당 모듈은 일~월 기준인데 우리는 토~일 기준이므로 1일을 더해줘서 월을 계산한다.
        sku_week_tmp = datetime.datetime(year,month,day) + timedelta(1)
        sku_week=sku_week_tmp.isocalendar()[1]
        #입력한 주차에 해당하는 애들만 본다
        if  '~' in answer:
            if int(start_date)>int(str(int(df_sandbox.iloc[i,7]))[0:8]) or int(end_date)<int(str(int(df_sandbox.iloc[i,7]))[0:8]):
                continue
        else:
            if int(now_year)!=int(sku_week_tmp.isocalendar()[0]) or int(now_week)!= int(sku_week):
                continue

        skuid = str(df_sandbox.iloc[i,0])
        skuname = str(df_sandbox.iloc[i,1])
        manufacture = str(df_sandbox.iloc[i,2])
        system_ulr_manu = str(df_sandbox.iloc[i,3])
        brand_kr =str(df_sandbox.iloc[i,4])
        if manufacture=='\x05':
            manufacture=''
        if brand_kr=='\x05':
            brand_kr=''
        system_ulr_brand=str(df_sandbox.iloc[i,5])
        audit_url=str(df_sandbox.iloc[i,6])
        enddt=str(int(df_sandbox.iloc[i,7]))
        status=str(df_sandbox.iloc[i,8])
        cate=str(df_sandbox.iloc[i,9])

        if status in status_list:
            continue

        if status in status_list:
            audit_complete='X'
        else:
            audit_complete='O'

        #print (skuid+"     "+cate)
        data_list[0][total_sku]=data_list[0][total_sku]+1
        data_list[cate_list.index(cate)][total_sku]=data_list[cate_list.index(cate)][total_sku]+1
        after_parsing_audit_url=""
        if is_valid_url(audit_url) :
            #print(str(df_sandbox.iloc[i]))
            after_parsing_audit_url=url_parsing(url_rear_part,audit_url)

        #11-Garbage 12-sys_url존재여부 13-op_url존재여부 14-일치여부
        sku_data=[enddt,skuid,skuname,manufacture,system_ulr_manu,brand_kr,system_ulr_brand,after_parsing_audit_url,audit_url,status,cate,'','','','']

        #garbage
        sku_data[11]=1
        #쓰레기값 빼기
        for keword in trash:
            if keword==manufacture:
                check_valid_manufacture=0
                sku_data[11]=3
                #제조사컬럼이 쓰레기가 들어왔을 경우
                continue
        for keword in trash:
            if keword==brand_kr:
                check_valid_brand_kr=0
                sku_data[11]=2
                #브랜드컬럼이 쓰레기가 들어왔을 경우
                continue
        if check_valid_manufacture==0 and check_valid_brand_kr==0:
            sku_data[11]=4

        #sys_url
        if is_valid_url(system_ulr_manu) and is_valid_url(system_ulr_brand):
            sku_data[12]=1
        if is_valid_url(system_ulr_manu) and (not is_valid_url(system_ulr_brand)):
            sku_data[12]=2
        if (not is_valid_url(system_ulr_manu)) and is_valid_url(system_ulr_brand):
            sku_data[12]=3
        if (not is_valid_url(system_ulr_manu)) and (not is_valid_url(system_ulr_brand)):
            sku_data[12]=4
            
        #op_url
        if is_valid_url(audit_url):
            sku_data[13]='O'
        else :
            sku_data[13]='X'

        #일치여부
        if is_valid_url(audit_url):
            if(after_parsing_audit_url==system_ulr_manu) or (after_parsing_audit_url==system_ulr_brand):
                sku_data[14]='O'
            else:
                sku_data[14]='X'
        else:
            sku_data[14]='X'

        #스큐 커버율,정확도 모수
        if sku_data[11]==4: #둘다 쓰레기일때 스큐 자체가 가비지인 경우
            data_list[total_sku][sku_trash]=data_list[total_sku][sku_trash]+1
            data_list[cate_list.index(cate)][sku_trash]=data_list[cate_list.index(cate)][sku_trash]+1
        else: #가비지가 아닌경우
            data_list[total_sku][sku_cover_mother]=data_list[total_sku][sku_cover_mother]+1
            data_list[cate_list.index(cate)][sku_cover_mother]=data_list[cate_list.index(cate)][sku_cover_mother]+1
            if sku_data[12]!=4 : #url이 한개라도 있는 경우
                data_list[total_sku][url_m_or_b]=data_list[total_sku][url_m_or_b]+1
                data_list[cate_list.index(cate)][url_m_or_b]=data_list[cate_list.index(cate)][url_m_or_b]+1
            else: #url이 둘다 없는 경우
                if sku_data[13]=='X' and audit_complete =='O': #오퍼레이터 url이 존재하지 않을때
                    data_list[total_sku][sku_cover_null_null]=data_list[total_sku][sku_cover_null_null]+1
                    data_list[cate_list.index(cate)][sku_cover_null_null]=data_list[cate_list.index(cate)][sku_cover_null_null]+1
            #작업을 하였을때
            if audit_complete =='O':
                data_list[total_sku][sku_precision_mother]=data_list[total_sku][sku_precision_mother]+1
                data_list[cate_list.index(cate)][sku_precision_mother]=data_list[cate_list.index(cate)][sku_precision_mother]+1
                #오퍼레이터 url이 있을때
                if sku_data[13]=='O' and sku_data[14]=='O':
                    data_list[total_sku][sku_hit]=data_list[total_sku][sku_hit]+1
                    data_list[cate_list.index(cate)][sku_hit]=data_list[cate_list.index(cate)][sku_hit]+1
                elif sku_data[12]==4 and sku_data[13]=='X':
                    data_list[total_sku][sku_precision_null_null]=data_list[total_sku][sku_precision_null_null]+1
                    data_list[cate_list.index(cate)][sku_precision_null_null]=data_list[cate_list.index(cate)][sku_precision_null_null]+1
        
        #제조사 커버율,정확도 모수
        if sku_data[11]==3 or sku_data[11]==4:
            data_list[total_sku][m_trash]=data_list[total_sku][m_trash]+1
            data_list[cate_list.index(cate)][m_trash]=data_list[cate_list.index(cate)][m_trash]+1
        else:
            data_list[total_sku][m_cover_mother]=data_list[total_sku][m_cover_mother]+1
            data_list[cate_list.index(cate)][m_cover_mother]=data_list[cate_list.index(cate)][m_cover_mother]+1
            if sku_data[12]==1 or  sku_data[12]==2:
                data_list[total_sku][url_m]=data_list[total_sku][url_m]+1
                data_list[cate_list.index(cate)][url_m]=data_list[cate_list.index(cate)][url_m]+1
                if(sku_data[14]=='O'):
                    data_list[total_sku][m_hit]=data_list[total_sku][m_hit]+1
                    data_list[cate_list.index(cate)][m_hit]=data_list[cate_list.index(cate)][m_hit]+1
            else:
                if sku_data[13]=='X' and audit_complete =='O':
                    data_list[total_sku][m_cover_null_null]=data_list[total_sku][m_cover_null_null]+1
                    data_list[cate_list.index(cate)][m_cover_null_null]=data_list[cate_list.index(cate)][m_cover_null_null]+1


        #브랜드 커버율,정확도 모수
        if sku_data[11]==2 or sku_data[11]==4 :
             data_list[total_sku][b_trash]=data_list[total_sku][b_trash]+1
             data_list[cate_list.index(cate)][b_trash]=data_list[cate_list.index(cate)][b_trash]+1
        else:
            data_list[total_sku][b_cover_mother]=data_list[total_sku][b_cover_mother]+1
            data_list[cate_list.index(cate)][b_cover_mother]=data_list[cate_list.index(cate)][b_cover_mother]+1
            if sku_data[12]==1 or  sku_data[12]==3:
                data_list[total_sku][url_b]=data_list[total_sku][url_b]+1
                data_list[cate_list.index(cate)][url_b]=data_list[cate_list.index(cate)][url_b]+1
                if(sku_data[14]=='O'):
                data_list[total_sku][b_hit]=data_list[total_sku][b_hit]+1
                data_list[cate_list.index(cate)][b_hit]=data_list[cate_list.index(cate)][b_hit]+1  
            else:
                if sku_data[13]=='X' and audit_complete =='O':
                    data_list[total_sku][b_cover_null_null]=data_list[total_sku][b_cover_null_null]+1
                    data_list[cate_list.index(cate)][b_cover_null_null]=data_list[cate_list.index(cate)][b_cover_null_null]+1

        ws_sku_list.append(sku_data) 
    now_line=1
    for t in range(len(cate_list)):
        ws.cell(row = now_line,column=t+2,value=str(cate_list[t]))
    now_line=now_line+1
    for y in range(len(list_of_var)):
        ws.cell(row = now_line,column=1,value=str(list_of_var[y]))
        for u in range(len(cate_list)):
            ws.cell(row = now_line,column=u+2,value=str(data_list[u][y]))                
        now_line=now_line+1

    for j in range(len(ws_sku_list)):
        for k in range(len(list_of_columm)):
            ws_sku.cell(row = sku_now_line,column=k+1,value=str(ws_sku_list[j][k]))
        sku_now_line=sku_now_line+1

    
    wb.save('./cover_precision_final' + answer + '.xlsx')

    

except Exception as e:
        print(e)